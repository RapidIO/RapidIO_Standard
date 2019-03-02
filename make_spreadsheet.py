#!/usr/bin/env python
# -*- coding: latin-1 -*-
"""
    Read in a text file and create an XLSX spreadsheet.

    Assumes that the first line of the file that contains a
    comma is a header line.

    Strips off leading and ending "'" delimiters for items in each line,
    if they exist.
"""

from optparse import OptionParser
from collections import OrderedDict
import operator
import re
import sys
import os
import logging
import copy
from difflib import Differ
from constants import *
from create_translation import *
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import codecs
import subprocess

class ExcelEditor(object):
    def __init__(self, text, excel, source = "text"):
        self.text_filepath = text
        self.excel_filepath = excel
        self.header = None
        self.data = []
        self.lines = []
        if source == "text":
            self._read_text()
            self._create_excel()
            self._format_excel()
        else:
            self._read_excel()

    def _stripped_tokens(self, line):
        toks = []
        if line[0] == "'" and line[-1] == "'":
            toks = [tok.strip() for tok in line[1:-1].split("', '")]
        else:
            toks = [tok.strip() for tok in line.split(",")]
        return toks

    def _read_text(self):
        logging.info("Reading text file '%s'." % self.text_filepath)
        with open(self.text_filepath, 'r') as text_file:
            self.lines = text_file.readlines()
        for idx, line in enumerate(self.lines):
            logging.info("Text %d: '%s'" % (idx, line))
        col_warning = False

        for l in self.lines:
            line = l.strip()
            if self.header is None:
                if line.find(',') >= 0:
                    self.header = self._stripped_tokens(line)
                continue

            toks = self._stripped_tokens(line)
            if len(toks) != len(self.header) and not col_warning:
                 logging.warning("Header different length than data: %d %d. Example:" % (len(self.header), len(toks)))
                 logging.warning("Header: %s" % self.header)
                 logging.warning("Data  : %s" % toks)
                 raise ValueError("Halting for debug...")
                 
            self.data.append(toks)

        if self.header is None or self.header == []:
            raise ValueError("No header found in file %s" % self.text_filepath)

    def _create_excel(self):
        self.wb = Workbook()
        ws = self.wb.active
        text_file_name = os.path.basename(self.text_filepath)
        ws.title = ''.join(x for x in text_file_name.title() if not x.isspace())

        header_cell_alignment = Alignment(horizontal="center", vertical="top",
                                          wrap_text=True, shrink_to_fit=True)

        for col, h_tok in enumerate(self.header):
            cell = ws.cell(row=1, column=col+1, value=h_tok)
            cell.alignment = header_cell_alignment

        data_cell_alignment = Alignment(horizontal="left", vertical="top",
                                        wrap_text=True, shrink_to_fit=True)
         
        for r, d_toks in enumerate(self.data):
            for c, d_tok in enumerate(d_toks):
                val = d_tok.decode("utf-8","ignore")
                a_val = val.encode("ascii","ignore")
                if a_val != '':
                    if (a_val[0] == "'"):
                        raise ValueError("Tokens &%s& token &%s& bad &%s&." %
                                              ("&".join(d_toks), d_tok, a_val))
                    a_val = a_val.replace('\\n', '\n')
                cell = ws.cell(row=r+2, column=c+1, value=str(a_val))
                cell.alignment = data_cell_alignment
        
    def _format_excel(self):
        ws = self.wb.active
        max_col_width = 60
        for column_cells in ws.columns:
            val_length = max(len(cell.value) for cell in column_cells)
            col_width = min(val_length, max_col_width)
            ws.column_dimensions[column_cells[0].column].width = col_width

    def _read_excel(self):
        self.wb = load_workbook(filename = self.excel_filepath)
        for sheet in self.wb:
            if sheet.title != self.wb.active.title:
                continue
            idx = 0
            for row in sheet.iter_rows(min_row = 1):
               vals = [c.value  for c in row]
               for i, val in enumerate(vals):
                   if val is None:
                       vals[i] = ''
                   try:
                       vals[i] = str(vals[i])
                   except UnicodeEncodeError:
                       vals[i].encode("ascii","ignore")
               line = "', '".join(vals)
               line = line.replace('\n', '\\n')
               line_crlf = "'%s'\n" % line
               self.lines.append(line_crlf)
               logging.info("Line from Excel %d: '%s'" % (idx, line_crlf))
               idx += 1

    def write_excel(self):
        self.wb.save(self.excel_filepath)

    def write_text(self):
        with open(self.text_filepath, 'w') as t:
            for l in self.lines:
                t.write(l)

def create_parser():
    parser = OptionParser(description="Create Excel spreadsheet based on text file.")
    parser.add_option('-t', '--textfile',
            dest = 'text_filepath',
            action = 'store', type = 'string', default = None,
            help = 'File path to text file.',
            metavar = 'FILE')
    parser.add_option('-x', '--excel',
            dest = 'excel_filepath',
            action = 'store', type = 'string', default = None,
            help = 'File path to new Excel file.',
            metavar = 'FILE')
    return parser

def validate_options(options):
    if not os.path.isfile(options.text_filepath):
        raise ValueError("Text file '%s' does not exist." %
                         options.text_filepath)
    if os.path.isfile(options.excel_filepath):
        raise ValueError("File '%s' will be overwritten!" %
                         options.excel_filepath)

def main(argv = None):
    logging.basicConfig(level=logging.WARN)
    parser = create_parser()
    if argv is None:
        argv = sys.argv[1:]

    (options, argv) = parser.parse_args(argv)
    if len(argv) != 0:
        print('Invalid argument!')
        print
        parser.print_help()
        return -1

    try:
        validate_options(options)
    except ValueError as e:
        print(e)
        sys.exit(-1)

    excel = ExcelEditor(options.text_filepath, options.excel_filepath)
    excel.write_excel()
    cmd = "exo-open %s" % options.excel_filepath
    process = subprocess.Popen(cmd.split(), stdout=subprocess.PIPE)
    output, error = process.communicate()
    inp = raw_input("Press enter when you've closed the spreadsheet.")
    new_excel = ExcelEditor(options.text_filepath, options.excel_filepath, "XL")
    new_excel.write_text()

if __name__ == '__main__':
    sys.exit(main())
